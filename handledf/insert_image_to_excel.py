


import openpyxl
from openpyxl.drawing.image import Image
import os
from PIL import Image as PILImage

def create_excel_with_images(excel_path, img_path, img_name_column='A', img_column='B'):
    print(f"Creating new Excel file: {excel_path}")
    print(f"Image directory: {img_path}")

    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Images"

    # Populate Excel with image names
    print("Populating Excel with image names")
    ws[f'{img_name_column}1'] = "Image Name"
    row = 2
    temp_files = []  # List to keep track of temporary files

    for filename in os.listdir(img_path):
        name, ext = os.path.splitext(filename)
        if ext.lower() in ['.jpg', '.jpeg', '.png', '.gif']:
            ws[f'{img_name_column}{row}'] = name
            print(f"Added image name to {img_name_column}{row}: {name}")

            # Insert image
            img_file_path = os.path.join(img_path, filename)
            try:
                # Open and resize the image
                with PILImage.open(img_file_path) as img:
                    img.thumbnail((100, 100))  # Resize image to 100x100 pixels
                    temp_path = f"temp_{name}.png"
                    img.save(temp_path)
                    temp_files.append(temp_path)  # Add to list of temporary files

                # Add image to Excel
                img = Image(temp_path)
                ws.add_image(img, f"{img_column}{row}")

                # Adjust column width and row height
                ws.column_dimensions[img_column].width = 15
                ws.row_dimensions[row].height = 75

                print(f"Successfully inserted image for {name}")
            except Exception as e:
                print(f"Error processing image {name}: {str(e)}")

            row += 1

    # Save the Excel file
    try:
        wb.save(excel_path)
        print(f"Excel file created and saved successfully at {excel_path}")
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")
    finally:
        wb.close()

    # Remove temporary files after saving the workbook
    for temp_file in temp_files:
        try:
            os.remove(temp_file)
            print(f"Removed temporary file: {temp_file}")
        except Exception as e:
            print(f"Error removing temporary file {temp_file}: {str(e)}")

if __name__ == "__main__":
    excel_path = './image.xlsx'
    img_path = './pic'
    
    create_excel_with_images(excel_path, img_path)

print("Script execution completed. Please check the console output for details.")



